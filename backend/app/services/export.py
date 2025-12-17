from typing import List
# import pandas as pd # Removed optimization
from io import BytesIO
from fpdf import FPDF
from datetime import date
from app.models import Equipamento

class PDF(FPDF):
    def header(self):
        self.set_font('Helvetica', 'B', 12)
        self.cell(0, 10, 'Relatório de Equipamentos - CalibraCore Lab', 0, new_x="LMARGIN", new_y="NEXT", align='C')
        self.ln(5)

    def footer(self):
        self.set_y(-15)
        self.set_font('Helvetica', 'I', 8)
        self.cell(0, 10, f'Página {self.page_no()}/{{nb}}', 0, new_x="LMARGIN", new_y="NEXT", align='C')


def format_status(status_value: str) -> str:
    status_map = {
        'em_dia': 'Em dia',
        'proximo_60': 'Vence em 60d',
        'proximo_30': 'Vence em 30d',
        'vencido': 'Vencido'
    }
    return status_map.get(status_value, status_value)

def format_date(date_obj) -> str:
    if not date_obj:
        return '-'
    return date_obj.strftime("%d/%m/%Y")

import openpyxl
from openpyxl.utils import get_column_letter

def generate_excel_report(equipamentos: List[Equipamento]) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Equipamentos"
    
    # Headers
    headers = [
        "Código Interno", "Descrição", "Categoria", "Marca", 
        "Nº Certificado", "Nº Série", "Laboratório", 
        "Última Calibração", "Vencimento", "Status", "Observações"
    ]
    ws.append(headers)
    
    # Data
    for eq in equipamentos:
        ws.append([
            eq.codigo_interno,
            eq.descricao,
            eq.categoria,
            eq.marca,
            eq.numero_certificado,
            eq.numero_serie,
            eq.laboratorio,
            format_date(eq.data_ultima_calibracao),
            format_date(eq.data_vencimento),
            format_status(eq.status.value),
            eq.observacoes
        ])
    
    # Auto-adjust columns width
    for i, col in enumerate(headers):
        max_len = len(col) + 2
        # Check content length in this column (simple check)
        column_letter = get_column_letter(i + 1)
        
        # Start checking from row 2 (data)
        # Note: Iterating all rows might be slow for massive data, but fine here
        # For simplicity/speed we can skip deep content analysis or limit it
        # But let's keep it simple
        ws.column_dimensions[column_letter].width = min(max_len + 20, 50) # Generic width + buffer

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def generate_pdf_report(equipamentos: List[Equipamento]) -> BytesIO:
    pdf = PDF(orientation='L', unit='mm', format='A4')
    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.set_font('Helvetica', '', 10)
    
    # Headers
    headers = ["Cód.", "Descrição", "Categoria", "Lab", "Vencimento", "Status"]
    # Adjusted widths to fit A4 Landscape (approx 277mm usable width)
    # Total: 45 + 85 + 40 + 35 + 30 + 35 = 270mm
    col_widths = [45, 85, 40, 35, 30, 35]
    
    pdf.set_fill_color(240, 240, 240)
    pdf.set_font('Helvetica', 'B', 10)
    for i, header in enumerate(headers):
        pdf.cell(col_widths[i], 10, header, border=1, align='C', fill=True)
    pdf.ln()
    
    # Data
    pdf.set_font('Helvetica', '', 9)
    for eq in equipamentos:
        # Check if page break is needed
        if pdf.get_y() > 180:
            pdf.add_page()
            # Re-print headers
            pdf.set_font('Helvetica', 'B', 10)
            for i, header in enumerate(headers):
                pdf.cell(col_widths[i], 10, header, border=1, align='C', fill=True)
            pdf.ln()
            pdf.set_font('Helvetica', '', 9)
            
        # Truncate text to fit cell width (approximate char width calculation)
        # Assuming avg char width is ~2mm for size 9 font
        def truncate(text, width_mm):
            if not text: return ""
            text = str(text)
            max_chars = int(width_mm / 2.2) # Conservative estimate
            if len(text) > max_chars:
                return text[:max_chars-2] + ".."
            return text

        status_text = format_status(eq.status.value)
        
        # Determine row height based on content? No, fixed height for simplicity in table view
        # Using truncate to ensure single line
        row_height = 8
        
        pdf.cell(col_widths[0], row_height, truncate(eq.codigo_interno, col_widths[0]), border=1)
        pdf.cell(col_widths[1], row_height, truncate(eq.descricao, col_widths[1]), border=1)
        pdf.cell(col_widths[2], row_height, truncate(eq.categoria, col_widths[2]), border=1)
        pdf.cell(col_widths[3], row_height, truncate(eq.laboratorio, col_widths[3]), border=1)
        pdf.cell(col_widths[4], row_height, format_date(eq.data_vencimento), border=1, align='C')
        
        # Color code status text if needed, or just print
        pdf.cell(col_widths[5], row_height, status_text, border=1, align='C')
        pdf.ln()

    output = BytesIO()
    pdf_bytes = pdf.output()
    output.write(pdf_bytes)
    output.seek(0)
    return output
